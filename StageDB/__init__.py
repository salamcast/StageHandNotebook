from openpyxl import load_workbook
import os

global DB 
DB = os.path.join(os.path.dirname(__file__), 'DB.xlsx')

def load_test():
    wb = load_workbook(DB)
    #print(wb.sheetnames)
    page = wb['Power']
    for x in page.values:
        print(x)